#!/usr/bin/env python3
"""
Extract release-notes text from a PDF into a one-row-per-line Excel file,
preserving structural formatting (headings, bullets, sub-bullets).

Approach
--------
PDFs don't carry semantic tags like <h1> or <li> — headings and bullets are
just visual conventions (font size, boldness, indentation, leading glyphs).
So this script:

  1. Extracts each line of text along with its font size / bold flag / x0
     position, using pdfplumber (works line-by-line via `extract_words`).
  2. Classifies each line into a Level (H1, H2, BULLET, SUB_BULLET, BODY)
     using calibrated thresholds — see `LineClassifier`.
  3. Writes one row per line into an .xlsx file via openpyxl, applying
     matching cell styling (bold, font size, indent) per Level.

Usage
-----
    python pdf_release_notes_to_xlsx.py input.pdf output.xlsx

Calibration
-----------
Real-world PDFs vary in how consistently they use font sizes for headings.
Run with `--inspect` first to print the raw (text, size, bold, x0) tuples
pdfplumber sees, then adjust the thresholds in `LineClassifier` to match
your document before doing the real extraction.

    python pdf_release_notes_to_xlsx.py input.pdf --inspect
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from enum import Enum
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

BULLET_PREFIX_RE = re.compile(r"^[\u2022\u2023\u25E6\u2043\u2219\-\*]\s+")
NUMBERED_PREFIX_RE = re.compile(r"^(\d+[\.\)]|\(\d+\))\s+")


class Level(str, Enum):
    H1 = "H1"
    H2 = "H2"
    H3 = "H3"
    BULLET = "Bullet"
    SUB_BULLET = "Sub-bullet"
    BODY = "Body"


@dataclass(frozen=True)
class RawLine:
    """A single line of text with the layout metadata pdfplumber exposes."""

    text: str
    font_size: float
    is_bold: bool
    x0: float
    page_number: int


@dataclass(frozen=True)
class ClassifiedLine:
    level: Level
    text: str
    page_number: int


class LineExtractor:
    """Groups pdfplumber's word-level output into lines with metadata.

    pdfplumber gives font info per *word*, not per line, so words are
    bucketed by their vertical position (`top`) and merged. The dominant
    font size/weight per line drives classification; small inline artifacts
    (e.g. a superscript footnote marker) shouldn't skew the whole line.
    """

    # Words within this many points of each other vertically are one line.
    LINE_TOLERANCE = 3.0

    def extract(self, pdf_path: Path) -> list[RawLine]:
        lines: list[RawLine] = []
        with pdfplumber.open(pdf_path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                words = page.extract_words(
                    extra_attrs=["size", "fontname"], use_text_flow=False
                )
                lines.extend(self._group_into_lines(words, page_number))
        return lines

    def _group_into_lines(self, words: list[dict], page_number: int) -> list[RawLine]:
        if not words:
            return []

        words = sorted(words, key=lambda w: (w["top"], w["x0"]))
        buckets: list[list[dict]] = []
        for word in words:
            if buckets and abs(word["top"] - buckets[-1][-1]["top"]) <= self.LINE_TOLERANCE:
                buckets[-1].append(word)
            else:
                buckets.append([word])

        raw_lines = []
        for bucket in buckets:
            bucket.sort(key=lambda w: w["x0"])
            text = " ".join(w["text"] for w in bucket).strip()
            if not text:
                continue
            sizes = [w["size"] for w in bucket]
            dominant_size = max(set(sizes), key=sizes.count)
            is_bold = any("bold" in w["fontname"].lower() for w in bucket)
            x0 = bucket[0]["x0"]
            raw_lines.append(RawLine(text, dominant_size, is_bold, x0, page_number))
        return raw_lines


class LineClassifier:
    """Maps raw layout metadata to a semantic Level.

    Thresholds are calibrated against the sizes actually present in the
    document rather than hardcoded absolutes, since "18pt is a heading"
    only holds for some PDFs. Run --inspect to see your document's actual
    size distribution and adjust BODY_SIZE_TOLERANCE / indent steps if the
    heuristic misclassifies anything.
    """

    BODY_SIZE_TOLERANCE = 0.5  # points; treat sizes within this as "same"
    INDENT_STEP = 15.0  # points of x0 movement that count as one indent level

    def __init__(self, raw_lines: list[RawLine]):
        sizes = [line.font_size for line in raw_lines]
        self._body_size = self._infer_body_size(sizes)
        self._base_x0 = min((line.x0 for line in raw_lines), default=0.0)

    @staticmethod
    def _infer_body_size(sizes: list[float]) -> float:
        """Body text is normally the most frequent font size in the doc."""
        if not sizes:
            return 10.0
        return max(set(sizes), key=sizes.count)

    def classify(self, line: RawLine) -> ClassifiedLine:
        text = line.text
        size_delta = line.font_size - self._body_size
        indent_level = round((line.x0 - self._base_x0) / self.INDENT_STEP)

        has_bullet_prefix = bool(BULLET_PREFIX_RE.match(text) or NUMBERED_PREFIX_RE.match(text))
        cleaned_text = BULLET_PREFIX_RE.sub("", text)
        cleaned_text = NUMBERED_PREFIX_RE.sub("", cleaned_text)

        if size_delta > 4 and line.is_bold:
            level = Level.H1
        elif size_delta > 1.5 and line.is_bold:
            level = Level.H2
        elif line.is_bold and abs(size_delta) <= self.BODY_SIZE_TOLERANCE:
            level = Level.H3
        elif has_bullet_prefix or indent_level >= 2:
            level = Level.SUB_BULLET if indent_level >= 2 else Level.BULLET
        else:
            level = Level.BODY

        return ClassifiedLine(level=level, text=cleaned_text.strip(), page_number=line.page_number)


class ReleaseNotesWorkbookWriter:
    """Writes classified lines to a one-row-per-line, styled .xlsx file."""

    HEADER_ROW = ("Page", "Level", "Text")

    STYLE_BY_LEVEL: dict[Level, dict] = {
        Level.H1: {"bold": True, "size": 16, "indent": 0},
        Level.H2: {"bold": True, "size": 13, "indent": 0},
        Level.H3: {"bold": True, "size": 11, "indent": 0},
        Level.BULLET: {"bold": False, "size": 11, "indent": 1},
        Level.SUB_BULLET: {"bold": False, "size": 11, "indent": 2},
        Level.BODY: {"bold": False, "size": 11, "indent": 0},
    }

    def write(self, lines: list[ClassifiedLine], output_path: Path) -> None:
        workbook = Workbook()
        sheet: Worksheet = workbook.active
        sheet.title = "Release Notes"
        sheet.append(self.HEADER_ROW)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for line in lines:
            style = self.STYLE_BY_LEVEL[line.level]
            row_index = sheet.max_row + 1
            sheet.append((line.page_number, line.level.value, line.text))

            text_cell = sheet.cell(row=row_index, column=3)
            text_cell.font = Font(bold=style["bold"], size=style["size"])
            text_cell.alignment = Alignment(
                indent=style["indent"] * 2, wrap_text=True, vertical="top"
            )

        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 100
        sheet.freeze_panes = "A2"

        workbook.save(output_path)


def inspect(pdf_path: Path) -> None:
    """Print raw (text, size, bold, x0) tuples to help calibrate thresholds."""
    extractor = LineExtractor()
    for line in extractor.extract(pdf_path):
        print(
            f"p{line.page_number:>3} | size={line.font_size:5.1f} | "
            f"bold={line.is_bold!s:5} | x0={line.x0:6.1f} | {line.text}"
        )


def convert(pdf_path: Path, output_path: Path) -> None:
    extractor = LineExtractor()
    raw_lines = extractor.extract(pdf_path)
    if not raw_lines:
        print("No text extracted — is this a scanned/image PDF? Try OCR first.", file=sys.stderr)
        sys.exit(1)

    classifier = LineClassifier(raw_lines)
    classified = [classifier.classify(line) for line in raw_lines]

    ReleaseNotesWorkbookWriter().write(classified, output_path)
    print(f"Wrote {len(classified)} lines to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("pdf_path", type=Path)
    parser.add_argument("xlsx_path", type=Path, nargs="?", help="Output .xlsx path")
    parser.add_argument("--inspect", action="store_true", help="Print raw layout metadata instead of converting")
    args = parser.parse_args()

    if args.inspect:
        inspect(args.pdf_path)
        return

    if args.xlsx_path is None:
        parser.error("xlsx_path is required unless --inspect is passed")

    convert(args.pdf_path, args.xlsx_path)


if __name__ == "__main__":
    main()